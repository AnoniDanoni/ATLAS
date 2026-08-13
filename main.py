# -*- coding: utf-8 -*-
# main.py - Aplicação principal ATLAS

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
import copy
from datetime import datetime

import set as set_module
from loading_screen import tela_carregamento

from core import (
    DIRETORIO_ATLAS, CONFIG_FILE,
    carregar_config, salvar_config, 
    obter_pastas_sessao, atualizar_pastas_sessao,
    detectar_versoes_revit, abrir_revit,
    verificar_atualizacoes, extrair_nome_base
)

from ui import (
    JanelaSelecaoRevit, JanelaSelecaoPastas,
    JanelaGerenciarSessoes, JanelaGerenciarPastas
)


# ==================== JANELA DE RELATÓRIO DE ARQUIVO INAPTO ====================

class JanelaRelatorioArquivo(tk.Toplevel):
    """Janela para criar/editar relatório de arquivo inapto"""
    
    def __init__(self, parent, nome_arquivo, caminho_arquivo, config, bg_principal, bg_secundario, fg_texto, fg_texto_secundario, btn_relatorio=None, combo=None):
        super().__init__(parent)
        self.title(f"📝 Relatório - {nome_arquivo}")
        self.geometry("600x500")
        self.configure(bg=bg_principal)
        self.grab_set()
        self.transient(parent)
        
        self.nome_arquivo = nome_arquivo
        self.caminho_arquivo = caminho_arquivo
        self.config = config
        self.btn_relatorio = btn_relatorio
        self.combo = combo
        self.foi_salvo = False
        self.bg_principal = bg_principal
        self.bg_secundario = bg_secundario
        self.bg_terciario = "#282B30"
        self.fg_texto = fg_texto
        self.fg_texto_secundario = fg_texto_secundario
        self.cor_acento = "#5865F2"
        
        # Inicializar estrutura de relatórios se não existir
        if 'relatorios_inaptid' not in self.config:
            self.config['relatorios_inaptid'] = {}
        
        # Capturar evento de fechamento
        self.protocol("WM_DELETE_WINDOW", self._ao_fechar)
        
        self._montar_interface()
        self._carregar_relatorio()
        
    def _montar_interface(self):
        """Monta a interface da janela"""
        # Header
        header_frame = tk.Frame(self, bg=self.bg_secundario)
        header_frame.pack(fill="x", pady=(0, 10))
        
        ttk.Label(header_frame, text=f"📝 Relatório de Arquivo Inapto", 
                 font=("Segoe UI", 11, "bold"), background=self.bg_secundario, 
                 foreground=self.cor_acento).pack(pady=10)
        
        # Info do arquivo
        info_frame = tk.Frame(self, bg=self.bg_principal)
        info_frame.pack(fill="x", padx=15, pady=(0, 10))
        
        ttk.Label(info_frame, text=f"Arquivo: {self.nome_arquivo}", 
                 font=("Segoe UI", 9), background=self.bg_principal, 
                 foreground=self.fg_texto).pack(anchor="w")
        
        ttk.Label(info_frame, text=f"Caminho: {self.caminho_arquivo}", 
                 font=("Segoe UI", 8), background=self.bg_principal, 
                 foreground=self.fg_texto_secundario).pack(anchor="w", pady=(3, 0))
        
        # Texto do relatório
        ttk.Label(self, text="Motivo da inaptidão:", 
                 font=("Segoe UI", 10, "bold"), background=self.bg_principal, 
                 foreground=self.fg_texto).pack(anchor="w", padx=15, pady=(0, 3))
        
        frame_texto = tk.Frame(self, bg=self.cor_acento, height=200)
        frame_texto.pack(fill="x", padx=15, pady=(0, 10))
        frame_texto.pack_propagate(False)
        
        frame_texto_inner = tk.Frame(frame_texto, bg=self.bg_principal)
        frame_texto_inner.pack(fill="both", expand=True, padx=2, pady=2)
        
        scrollbar = ttk.Scrollbar(frame_texto_inner)
        scrollbar.pack(side="right", fill="y")
        
        self.texto_relatorio = tk.Text(frame_texto_inner, 
                                       bg=self.bg_terciario, 
                                       fg=self.fg_texto,
                                       font=("Segoe UI", 9),
                                       height=8,
                                       relief="flat",
                                       borderwidth=0,
                                       yscrollcommand=scrollbar.set,
                                       wrap="word",
                                       padx=8,
                                       pady=8)
        self.texto_relatorio.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self.texto_relatorio.yview)
        
        # Botões
        btn_frame = tk.Frame(self, bg=self.bg_principal)
        btn_frame.pack(fill="x", padx=15, pady=10)
        
        ttk.Button(btn_frame, text="Cancelar", command=self._ao_fechar, width=20).pack(side="right", padx=(5, 0))
        ttk.Button(btn_frame, text="✓ OK", command=self._salvar, width=20).pack(side="right")

    
    def _carregar_relatorio(self):
        """Carrega o relatório salvo se existir"""
        if self.caminho_arquivo in self.config['relatorios_inaptid']:
            relatorio = self.config['relatorios_inaptid'][self.caminho_arquivo]
            self.texto_relatorio.insert("1.0", relatorio)
    
    def _ao_fechar(self):
        """Handler para fechar a janela (sem salvar)"""
        if not self.foi_salvo:
            # Se não foi salvo e existe um combo, restaurar para "Status"
            if self.combo:
                self.combo.set("Status")
                # Limpar o status salvo também
                if 'status_arquivos' in self.config and self.caminho_arquivo in self.config['status_arquivos']:
                    del self.config['status_arquivos'][self.caminho_arquivo]
                    salvar_config(self.config)
        
        self.destroy()
    
    def _salvar(self):
        """Salva o relatório"""
        texto = self.texto_relatorio.get("1.0", "end-1c")
        
        if not texto.strip():
            messagebox.showwarning("Aviso", "Digite o motivo da inaptidão!")
            return
        
        if 'relatorios_inaptid' not in self.config:
            self.config['relatorios_inaptid'] = {}
        
        self.config['relatorios_inaptid'][self.caminho_arquivo] = texto

        salvar_config(self.config)
        
        # Mostrar o botão de relatório se foi passado
        if self.btn_relatorio and not self.btn_relatorio.winfo_ismapped():
            self.btn_relatorio.pack(side="left", padx=2, pady=3)
        
        self.foi_salvo = True
        messagebox.showinfo("Sucesso", "Relatório salvo com sucesso!")
        self.destroy()


# ==================== JANELA DE RELATÓRIO ====================

class JanelaRelatorio(tk.Frame):
    def __init__(self, parent, config, resultados, bg_principal, bg_secundario, fg_texto, fg_texto_secundario, 
                 filter_novos, filter_desatualizados, filter_atualizados, on_voltar=None):
        super().__init__(parent)
        self.configure(bg=bg_principal)
        
        self.parent = parent
        self.on_voltar = on_voltar
        self.config = config
        self.resultados = resultados
        self.bg_principal = bg_principal
        self.bg_secundario = bg_secundario
        self.bg_terciario = "#282B30"
        self.fg_texto = fg_texto
        self.fg_texto_secundario = fg_texto_secundario
        self.cor_acento = "#FFFFFF"
        self.cor_novo = "#4D8AC8"
        self.cor_desatualizado = "#954EF1"
        self.cor_atualizado = "#5865F2"
        
        self.filter_novos = filter_novos
        self.filter_desatualizados = filter_desatualizados
        self.filter_atualizados = filter_atualizados
        self.arquivos_selecionados = set()
        self.cards_status = {}
        self.ignorados_relatorio_selecionados = set()
        self.ignorados_relatorio_vars = {}
        
        self._montar_interface()
        self._aplicar_filtros()

    def _montar_interface(self):
        """Monta a interface da janela de relatório"""
        self.configure(bg="#151821")
        self.bg_relatorio = "#151821"
        self.bg_lateral = "#1C202B"
        self.bg_card = "#1D202B"
        self.bg_card_hover = "#252A38"
        self.bg_nav_ativa = "#272C3F"
        self.fg_relatorio = "#E8EAF2"
        self.fg_relatorio_secundario = "#9EA3B3"
        self.cor_novo = "#86A3FF"
        self.cor_desatualizado = "#C77DFF"
        self.cor_atualizado = "#8EA6FF"
        self.bg_principal = self.bg_relatorio
        self.bg_secundario = self.bg_lateral
        self.bg_terciario = self.bg_card
        self.fg_texto = self.fg_relatorio
        self.fg_texto_secundario = self.fg_relatorio_secundario

        style = ttk.Style()
        style.configure(
            "Relatorio.TCombobox",
            fieldbackground="#272C3F",
            background="#272C3F",
            foreground=self.fg_relatorio,
            padding=8
        )
        style.map(
            "Relatorio.TCombobox",
            fieldbackground=[("readonly", "#272C3F")],
            background=[("readonly", "#272C3F"), ("active", "#30364A")],
            foreground=[("readonly", self.fg_relatorio)]
        )

        layout = tk.Frame(self, bg=self.bg_relatorio)
        layout.pack(fill="both", expand=True)

        sidebar = tk.Frame(layout, bg=self.bg_lateral, width=190)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        self.nav_botoes = {}
        nav_specs = [
            ("todos", "Todos"),
            ("novos", "Novos"),
            ("desatualizados", "Desatualizados"),
            ("atualizados", "Atualizados"),
            ("ignorados", "Ignorados"),
        ]
        for chave, texto in nav_specs:
            btn = tk.Button(
                sidebar,
                text=texto,
                anchor="w",
                bg=self.bg_lateral,
                fg=self.fg_relatorio_secundario,
                activebackground=self.bg_nav_ativa,
                activeforeground=self.cor_novo,
                relief="flat",
                bd=0,
                padx=16,
                pady=12,
                font=("Segoe UI", 11),
                cursor="hand2",
                command=lambda c=chave: self._selecionar_aba(c)
            )
            btn.pack(fill="x", padx=14, pady=(32 if chave == "todos" else 0, 6))
            self.nav_botoes[chave] = btn

        tk.Frame(sidebar, bg=self.bg_lateral).pack(fill="both", expand=True)
        tk.Button(
            sidebar,
            text="Voltar",
            anchor="w",
            bg=self.bg_lateral,
            fg=self.fg_relatorio_secundario,
            activebackground=self.bg_nav_ativa,
            activeforeground=self.fg_relatorio,
            relief="flat",
            bd=0,
            padx=16,
            pady=12,
            font=("Segoe UI", 11),
            cursor="hand2",
            command=self._ao_fechar
        ).pack(fill="x", padx=14, pady=(0, 24))

        area = tk.Frame(layout, bg=self.bg_relatorio)
        area.pack(side="left", fill="both", expand=True, padx=32, pady=28)

        header_frame = tk.Frame(area, bg=self.bg_relatorio)
        header_frame.pack(fill="x", pady=(0, 22))

        tk.Label(
            header_frame,
            text="Relatório de atualizações",
            bg=self.bg_relatorio,
            fg=self.fg_relatorio,
            font=("Segoe UI", 16)
        ).pack(side="left")

        self.csv_icon = self._criar_icone_csv("#272C3F", self.fg_relatorio, self.cor_novo)
        btn_exportar = tk.Frame(
            header_frame,
            bg="#272C3F",
            cursor="hand2"
        )
        btn_exportar.pack(side="right")
        icon_label = tk.Label(btn_exportar, image=self.csv_icon, bg="#272C3F", cursor="hand2")
        icon_label.pack(side="left", padx=(12, 6), pady=9)
        text_label = tk.Label(
            btn_exportar,
            text="CSV",
            bg="#272C3F",
            fg=self.fg_relatorio,
            font=("Segoe UI", 11, "bold"),
            cursor="hand2"
        )
        text_label.pack(side="left", padx=(0, 12), pady=9)

        def exportar_enter(event):
            btn_exportar.config(bg="#30364A")
            icon_label.config(bg="#30364A")
            text_label.config(bg="#30364A")

        def exportar_leave(event):
            btn_exportar.config(bg="#272C3F")
            icon_label.config(bg="#272C3F")
            text_label.config(bg="#272C3F")

        for widget in (btn_exportar, icon_label, text_label):
            widget.bind("<Button-1>", lambda event: self._exportar_csv())
            widget.bind("<Enter>", exportar_enter)
            widget.bind("<Leave>", exportar_leave)

        resumo_frame = tk.Frame(area, bg=self.bg_relatorio)
        resumo_frame.pack(fill="x", pady=(0, 22))

        self.resumo_labels = {
            "novos": self._criar_badge(resumo_frame, "0 novos", "#1F2A61", self.cor_novo),
            "desatualizados": self._criar_badge(resumo_frame, "0 desatualizados", "#4C246B", self.cor_desatualizado),
            "atualizados": self._criar_badge(resumo_frame, "0 atualizados", "#1F2A61", self.cor_atualizado),
        }

        self.conteudo_stack = tk.Frame(area, bg=self.bg_relatorio)
        self.conteudo_stack.pack(fill="both", expand=True)

        self.tab_todos = tk.Frame(self.conteudo_stack, bg=self.bg_relatorio)
        self.tab_novos = tk.Frame(self.conteudo_stack, bg=self.bg_relatorio)
        self.tab_desatualizados = tk.Frame(self.conteudo_stack, bg=self.bg_relatorio)
        self.tab_atualizados = tk.Frame(self.conteudo_stack, bg=self.bg_relatorio)
        self.tab_ignorados = tk.Frame(self.conteudo_stack, bg=self.bg_relatorio)

        for tab in (self.tab_todos, self.tab_novos, self.tab_desatualizados, self.tab_atualizados, self.tab_ignorados):
            tab.place(relx=0, rely=0, relwidth=1, relheight=1)
            self._criar_aba_scroll(tab)

        self._selecionar_aba("todos")

    def _criar_icone_csv(self, bg, fg, accent):
        icon = tk.PhotoImage(width=14, height=14)
        icon.put(bg, to=(0, 0, 14, 14))
        icon.put(fg, to=(2, 2, 10, 3))
        icon.put(fg, to=(2, 2, 3, 12))
        icon.put(fg, to=(2, 11, 12, 12))
        icon.put(fg, to=(11, 5, 12, 12))
        icon.put(fg, to=(9, 3, 10, 4))
        icon.put(fg, to=(10, 4, 11, 5))
        icon.put(accent, to=(4, 8, 10, 9))
        return icon

    def _criar_badge(self, parent, texto, bg, fg):
        badge = tk.Label(
            parent,
            text=texto,
            bg=bg,
            fg=fg,
            font=("Segoe UI", 10),
            padx=16,
            pady=5
        )
        badge.pack(side="left", padx=(0, 16))
        return badge

    def _selecionar_aba(self, chave):
        self.aba_ativa = chave
        tabs = {
            "todos": self.tab_todos,
            "novos": self.tab_novos,
            "desatualizados": self.tab_desatualizados,
            "atualizados": self.tab_atualizados,
            "ignorados": self.tab_ignorados,
        }
        if chave in tabs:
            tabs[chave].tkraise()

        for nav_chave, btn in self.nav_botoes.items():
            if nav_chave == chave:
                btn.config(bg=self.bg_nav_ativa, fg=self.cor_novo)
            else:
                btn.config(bg=self.bg_lateral, fg=self.fg_relatorio_secundario)

    def _titulo_pasta(self, caminho):
        nome = os.path.basename(os.path.normpath(caminho))
        return nome or caminho

    def _criar_aba_scroll(self, tab_frame):
        """Cria um widget canvas com scroll para uma aba com controles por arquivo"""
        canvas_frame = tk.Frame(tab_frame, bg=self.bg_relatorio)
        canvas_frame.pack(fill="both", expand=True)
        
        scrollbar = ttk.Scrollbar(canvas_frame)
        scrollbar.pack(side="right", fill="y")
        
        canvas = tk.Canvas(canvas_frame, bg=self.bg_relatorio, highlightthickness=0,
                          yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=canvas.yview)
        
        frame_conteudo = tk.Frame(canvas, bg=self.bg_relatorio)
        canvas_window = canvas.create_window((0, 0), window=frame_conteudo, anchor="nw")
        
        def atualizar_scroll_region(event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(canvas_window, width=canvas.winfo_width())
        
        def ao_scroll_mouse(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            return "break"
        
        def _bind_mousewheel(widget):
            """Bind mousewheel recursivamente em todos os widgets"""
            widget.bind("<MouseWheel>", ao_scroll_mouse, add="+")
            for child in widget.winfo_children():
                _bind_mousewheel(child)
        
        frame_conteudo.bind("<Configure>", atualizar_scroll_region)
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(canvas_window, width=e.width))
        
        # Bind mousewheel em todo o frame
        _bind_mousewheel(frame_conteudo)
        canvas.bind("<MouseWheel>", ao_scroll_mouse, add="+")
        tab_frame.bind("<MouseWheel>", ao_scroll_mouse, add="+")
        
        tab_frame._canvas = canvas
        tab_frame._frame_conteudo = frame_conteudo
        tab_frame._canvas_window = canvas_window
        tab_frame._bind_mousewheel = _bind_mousewheel

    def _criar_label_selecionavel(self, parent, texto, bg, fg, font_spec, padx=5, pady=3):
        """Cria um widget de texto selecionável sem bordas, estilizado como label"""
        text_widget = tk.Text(parent, height=1, width=max(20, len(texto)), bg=bg, fg=fg,
                             font=font_spec, relief="flat", borderwidth=0,
                             cursor="arrow", wrap="none", insertwidth=0)
        text_widget.insert("1.0", texto)
        text_widget.config(state="disabled")
        text_widget.pack(side="left", padx=padx, pady=pady)
        
        # O widget Text com state="disabled" permite seleção de texto e Ctrl+C automaticamente
        
        return text_widget

    def _aplicar_filtros(self):
        """Aplica filtros e atualiza o relatório"""
        self.arquivos_selecionados.clear()
        self.cards_status.clear()
        filtros = {
            'novos': self.filter_novos.get(),
            'desatualizados': self.filter_desatualizados.get(),
            'atualizados': self.filter_atualizados.get()
        }
        
        if not self.resultados:
            for widget in self.tab_todos._frame_conteudo.winfo_children():
                widget.destroy()
            lbl = tk.Label(self.tab_todos._frame_conteudo, text="Nenhum resultado disponível.\nExecute 'Verificar Atualizações' primeiro.",
                          bg=self.bg_relatorio, fg=self.fg_relatorio, font=("Segoe UI", 11))
            lbl.pack(pady=20)
            return

        total_novos = sum(len(r['novos']) for r in self.resultados.values())
        total_desatualizados = sum(len(r['desatualizados']) for r in self.resultados.values())
        total_atualizados = sum(len(r['atualizados']) for r in self.resultados.values())

        self.resumo_labels['novos'].config(text=f"{total_novos} novos")
        self.resumo_labels['desatualizados'].config(text=f"{total_desatualizados} desatualizados")
        self.resumo_labels['atualizados'].config(text=f"{total_atualizados} atualizados")

        self._preencher_aba_todos(filtros)
        self._preencher_aba_novos(filtros)
        self._preencher_aba_desatualizados(filtros)
        self._preencher_aba_atualizados(filtros)
        self._preencher_aba_ignorados()

        for tab in (self.tab_todos, self.tab_novos, self.tab_desatualizados, self.tab_atualizados, self.tab_ignorados):
            tab._bind_mousewheel(tab._frame_conteudo)

    def _preencher_aba_todos(self, filtros):
        """Preenche a aba de todos os resultados com dropdown para cada arquivo"""
        for widget in self.tab_todos._frame_conteudo.winfo_children():
            widget.destroy()
        
        total_novos = sum(len(r['novos']) for r in self.resultados.values())
        total_desatualizados = sum(len(r['desatualizados']) for r in self.resultados.values())
        total_atualizados = sum(len(r['atualizados']) for r in self.resultados.values())
        
        if total_novos == 0 and total_desatualizados == 0 and total_atualizados == 0:
            lbl = tk.Label(self.tab_todos._frame_conteudo, 
                          text="ℹ️ Nenhum resultado disponível.\nExecute 'Verificar Atualizações' primeiro.", 
                          bg=self.bg_principal, fg=self.fg_texto, font=("Segoe UI", 10))
            lbl.pack(pady=20)
            return
        
        if filtros['novos'] and total_novos > 0:
            titulo_frame = tk.Frame(self.tab_todos._frame_conteudo, bg=self.bg_principal)
            titulo_frame.pack(fill="x", padx=10, pady=(10, 2))
            tk.Label(titulo_frame, text=f"Novos ({total_novos})", bg=self.bg_principal,
                    fg=self.cor_novo, font=("Segoe UI", 9, "bold")).pack(anchor="w")
            
            for caminho, resultado in sorted(self.resultados.items(), key=lambda x: x[0].lower()):
                novos = resultado['novos']
                if not novos:
                    continue
                
                pasta_frame = tk.Frame(self.tab_todos._frame_conteudo, bg=self.bg_principal)
                pasta_frame.pack(fill="x", padx=10, pady=(5, 2))
                self._criar_label_selecionavel(pasta_frame, self._titulo_pasta(caminho), bg=self.bg_principal,
                        fg=self.cor_novo, font_spec=("Segoe UI", 10, "bold"))
                
                for item in novos:
                    nome_arquivo = f"{item['nome_base']}.{item['filtro_entrada']}"
                    caminho_arquivo = item['entrada']['caminho']
                    nome_base = item['nome_base']
                    self._adicionar_item_arquivo(self.tab_todos._frame_conteudo, nome_arquivo, caminho, caminho_arquivo, nome_base)
        
        if filtros['desatualizados'] and total_desatualizados > 0:
            titulo_frame = tk.Frame(self.tab_todos._frame_conteudo, bg=self.bg_principal)
            titulo_frame.pack(fill="x", padx=10, pady=(10, 2))
            tk.Label(titulo_frame, text=f"Desatualizados ({total_desatualizados})", bg=self.bg_principal,
                    fg=self.cor_desatualizado, font=("Segoe UI", 9, "bold")).pack(anchor="w")
            
            for caminho, resultado in sorted(self.resultados.items(), key=lambda x: x[0].lower()):
                desatualizados = resultado['desatualizados']
                if not desatualizados:
                    continue
                
                pasta_frame = tk.Frame(self.tab_todos._frame_conteudo, bg=self.bg_principal)
                pasta_frame.pack(fill="x", padx=10, pady=(5, 2))
                self._criar_label_selecionavel(pasta_frame, self._titulo_pasta(caminho), bg=self.bg_principal,
                        fg=self.cor_desatualizado, font_spec=("Segoe UI", 10, "bold"))
                
                for item in desatualizados:
                    nome_arquivo = f"{item['nome_base']} (há {item['dias']} dia(s))"
                    caminho_arquivo = item['entrada']['caminho']
                    nome_base = item['nome_base']
                    self._adicionar_item_arquivo(self.tab_todos._frame_conteudo, nome_arquivo, caminho, caminho_arquivo, nome_base)
        
        if filtros['atualizados'] and total_atualizados > 0:
            titulo_frame = tk.Frame(self.tab_todos._frame_conteudo, bg=self.bg_principal)
            titulo_frame.pack(fill="x", padx=10, pady=(10, 2))
            tk.Label(titulo_frame, text=f"Atualizados ({total_atualizados})", bg=self.bg_principal,
                    fg=self.cor_atualizado, font=("Segoe UI", 9, "bold")).pack(anchor="w")
            
            for caminho, resultado in sorted(self.resultados.items(), key=lambda x: x[0].lower()):
                atualizados = resultado['atualizados']
                if not atualizados:
                    continue
                
                pasta_frame = tk.Frame(self.tab_todos._frame_conteudo, bg=self.bg_principal)
                pasta_frame.pack(fill="x", padx=10, pady=(5, 2))
                self._criar_label_selecionavel(pasta_frame, self._titulo_pasta(caminho), bg=self.bg_principal,
                        fg=self.cor_atualizado, font_spec=("Segoe UI", 10, "bold"))
                
                for item in atualizados:
                    nome_arquivo = item['nome_base']
                    caminho_arquivo = item['entrada']['caminho']
                    nome_base = item['nome_base']
                    self._adicionar_item_arquivo(self.tab_todos._frame_conteudo, nome_arquivo, caminho, caminho_arquivo, nome_base)

    def _preencher_aba_novos(self, filtros):
        """Preenche a aba de novos con dropdown para cada arquivo"""
        for widget in self.tab_novos._frame_conteudo.winfo_children():
            widget.destroy()
        
        if not filtros['novos']:
            lbl = tk.Label(self.tab_novos._frame_conteudo, text="Filtro desativado", 
                          bg=self.bg_principal, fg=self.fg_texto, font=("Segoe UI", 9))
            lbl.pack(pady=20)
            return
        
        total_novos = sum(len(r['novos']) for r in self.resultados.values())
        
        if total_novos == 0:
            lbl = tk.Label(self.tab_novos._frame_conteudo, 
                          text="✅ Nenhum arquivo novo\n(Todos os arquivos já foram exportados)", 
                          bg=self.bg_principal, fg=self.fg_texto, font=("Segoe UI", 10))
            lbl.pack(pady=20)
            return
        
        titulo_frame = tk.Frame(self.tab_novos._frame_conteudo, bg=self.bg_principal)
        titulo_frame.pack(fill="x", padx=10, pady=(10, 5))
        tk.Label(titulo_frame, text=f"Total: {total_novos} arquivo(s) novo(s)",
                bg=self.bg_principal, fg=self.cor_novo, font=("Segoe UI", 9, "bold")).pack(anchor="w")
        
        for caminho, resultado in sorted(self.resultados.items(), key=lambda x: x[0].lower()):
            novos = resultado['novos']
            if not novos:
                continue
            
            pasta_frame = tk.Frame(self.tab_novos._frame_conteudo, bg=self.bg_principal)
            pasta_frame.pack(fill="x", padx=10, pady=(10, 2))
            self._criar_label_selecionavel(pasta_frame, self._titulo_pasta(caminho), bg=self.bg_principal,
                    fg=self.cor_novo, font_spec=("Segoe UI", 10, "bold"))
            
            config_frame = tk.Frame(self.tab_novos._frame_conteudo, bg=self.bg_principal)
            config_frame.pack(fill="x", padx=20, pady=(0, 5))
            tk.Label(config_frame, text=f"Filtro: {resultado['config']['entrada'].upper()} → {resultado['config']['saida'].upper()}", 
                    bg=self.bg_principal, fg=self.fg_texto_secundario, font=("Segoe UI", 8)).pack(anchor="w")
            
            for item in novos:
                nome_arquivo = f"{item['nome_base']}.{item['filtro_entrada']}"
                caminho_arquivo = item['entrada']['caminho']
                nome_base = item['nome_base']
                self._adicionar_item_arquivo(self.tab_novos._frame_conteudo, nome_arquivo, caminho, caminho_arquivo, nome_base)

    def _adicionar_item_arquivo(self, parent_frame, nome_arquivo, caminho_pasta, caminho_arquivo, nome_base=""):
        """Adiciona um item de arquivo com dropdown e botão de copiar caminho"""
        if not nome_base:
            nome_base = nome_arquivo.split('(')[0].strip() if '(' in nome_arquivo else nome_arquivo.split(' ')[0].strip()

        item_frame = tk.Frame(parent_frame, bg=self.bg_card)
        item_frame.pack(fill="x", padx=0, pady=(0, 8), ipady=9)

        nome_label = tk.Label(
            item_frame,
            text=nome_arquivo,
            bg=self.bg_card,
            fg=self.fg_relatorio,
            font=("Segoe UI", 10, "bold"),
            anchor="w"
        )
        nome_label.pack(side="left", fill="x", expand=True, padx=18)

        acoes_frame = tk.Frame(item_frame, bg=self.bg_card)
        acoes_frame.pack(side="right", padx=14)

        combo = ttk.Combobox(acoes_frame, values=["", "Atualizado", "Inapto", "Ignorar"],
                            state="readonly", width=11, font=("Segoe UI", 9),
                            style="Relatorio.TCombobox")

        if 'status_arquivos' in self.config and caminho_arquivo in self.config['status_arquivos']:
            combo.set(self.config['status_arquivos'][caminho_arquivo])
        else:
            combo.set("Status")

        combo.pack(side="left", padx=(0, 12), ipady=3)

        btn_copiar = tk.Button(acoes_frame, text="Copiar caminho", bg=self.bg_card,
                              fg=self.cor_novo, activebackground=self.bg_card_hover,
                              activeforeground=self.cor_novo, font=("Segoe UI", 9),
                              relief="flat", width=14, padx=8, pady=2, bd=0, cursor="hand2")
        btn_copiar.pack(side="left", padx=2)

        btn_relatorio = tk.Button(acoes_frame, text="relatório", bg=self.bg_card,
                                 fg="#2ECC71", activebackground=self.bg_card_hover,
                                 activeforeground="#2ECC71", font=("Segoe UI", 9),
                                 relief="flat", padx=6, pady=2, bd=0, cursor="hand2")
        if ('relatorios_inaptid' in self.config and caminho_arquivo in self.config['relatorios_inaptid'] and
            'status_arquivos' in self.config and self.config['status_arquivos'].get(caminho_arquivo) == 'Inapto'):
            btn_relatorio.pack(side="left", padx=(10, 0))

        card = {
            "frame": item_frame, "label": nome_label, "acoes": acoes_frame,
            "combo": combo, "btn_copiar": btn_copiar, "btn_relatorio": btn_relatorio,
            "pasta": caminho_pasta, "base": nome_base
        }
        self.cards_status.setdefault(caminho_arquivo, []).append(card)

        def alternar_selecao(event=None):
            selecionado = caminho_arquivo not in self.arquivos_selecionados
            (self.arquivos_selecionados.add if selecionado else self.arquivos_selecionados.discard)(caminho_arquivo)
            cor = self.bg_nav_ativa if selecionado else self.bg_card
            for item in self.cards_status.get(caminho_arquivo, []):
                item["frame"].config(bg=cor)
                item["label"].config(bg=cor)
                item["acoes"].config(bg=cor)
                item["btn_copiar"].config(bg=cor)
                item["btn_relatorio"].config(bg=cor)

        item_frame.bind("<Button-1>", alternar_selecao)
        nome_label.bind("<Button-1>", alternar_selecao)

        def ao_mudar_status(event=None):
            status = "Status" if combo.get() in ("", "Status") else combo.get()
            alvos = list(self.arquivos_selecionados) or [caminho_arquivo]

            if status == "Ignorar" and not messagebox.askyesno("Confirmar", f"Deseja ignorar {len(alvos)} arquivo(s)?"):
                combo.set(self.config.get('status_arquivos', {}).get(caminho_arquivo, "Status"))
                return

            for alvo in alvos:
                item = self.cards_status[alvo][0]
                if status == "Status":
                    self.config.get('status_arquivos', {}).pop(alvo, None)
                    self.config.get('relatorios_inaptid', {}).pop(alvo, None)
                elif status == "Ignorar":
                    self.config.setdefault('arquivos_ignorados', {}).setdefault(item["pasta"], [])
                    if item["base"] not in self.config['arquivos_ignorados'][item["pasta"]]:
                        self.config['arquivos_ignorados'][item["pasta"]].append(item["base"])
                    self.config.get('status_arquivos', {}).pop(alvo, None)
                    self.config.get('relatorios_inaptid', {}).pop(alvo, None)
                    self._remover_arquivo_ignorado(item["base"], item["pasta"])
                else:
                    anterior = self.config.get('status_arquivos', {}).get(alvo)
                    self.config.setdefault('status_arquivos', {})[alvo] = status
                    if status == "Inapto" and anterior and anterior != "Inapto":
                        self.config.get('relatorios_inaptid', {}).pop(alvo, None)

                for card_item in self.cards_status.get(alvo, []):
                    card_item["combo"].set(status)
                    if status == "Ignorar":
                        card_item["frame"].pack_forget()
                        self.arquivos_selecionados.discard(alvo)
                    elif status == "Inapto" and card_item["btn_relatorio"].winfo_ismapped() == 0:
                        card_item["btn_relatorio"].pack(side="left", padx=(10, 0))
                    elif status != "Inapto" and card_item["btn_relatorio"].winfo_ismapped():
                        card_item["btn_relatorio"].pack_forget()

                if status == "Ignorar":
                    self.cards_status.pop(alvo, None)

            salvar_config(self.config)
            if status == "Ignorar":
                self.after_idle(self._aplicar_filtros)

        def ao_clicar_copiar():
            self.clipboard_clear()
            self.clipboard_append(os.path.dirname(caminho_arquivo))
            self.update()
            btn_copiar.config(text="Copiado", fg="#2ECC71")

            def restaurar_botao_copiar():
                btn_copiar.config(text="Copiar caminho", fg=self.cor_novo)

            self.after(1800, restaurar_botao_copiar)

        def ao_clicar_relatorio():
            JanelaRelatorioArquivo(self, nome_arquivo, caminho_arquivo, self.config,
                                 self.bg_principal, self.bg_secundario, self.fg_texto,
                                 self.fg_texto_secundario, btn_relatorio, combo)

        combo.bind("<<ComboboxSelected>>", ao_mudar_status)
        btn_copiar.config(command=ao_clicar_copiar)
        btn_relatorio.config(command=ao_clicar_relatorio)

        def cor_card_atual():
            return self.bg_nav_ativa if caminho_arquivo in self.arquivos_selecionados else self.bg_card

        for btn in (btn_copiar, btn_relatorio):
            btn.bind("<Enter>", lambda event, b=btn: b.config(bg=self.bg_card_hover))
            btn.bind("<Leave>", lambda event, b=btn: b.config(bg=cor_card_atual()))

    def _remover_arquivo_ignorado(self, nome_base, caminho_pasta):
        """Remove o arquivo ignorado dos resultados"""
        if caminho_pasta not in self.resultados:
            return
        
        resultado = self.resultados[caminho_pasta]
        
        self.resultados[caminho_pasta]['novos'] = [
            r for r in resultado['novos'] 
            if r.get('nome_base', '') != nome_base
        ]
        
        self.resultados[caminho_pasta]['desatualizados'] = [
            r for r in resultado['desatualizados'] 
            if r.get('nome_base', '') != nome_base
        ]
        
        self.resultados[caminho_pasta]['atualizados'] = [
            r for r in resultado['atualizados'] 
            if r.get('nome_base', '') != nome_base
        ]

    def _preencher_aba_ignorados(self):
        for widget in self.tab_ignorados._frame_conteudo.winfo_children():
            widget.destroy()

        ignorados = {
            pasta: arquivos
            for pasta, arquivos in self.config.get('arquivos_ignorados', {}).items()
            if arquivos
        }
        total = sum(len(arquivos) for arquivos in ignorados.values())

        header = tk.Frame(self.tab_ignorados._frame_conteudo, bg=self.bg_principal)
        header.pack(fill="x", padx=10, pady=(10, 12))
        tk.Label(header, text=f"Ignorados ({total})", bg=self.bg_principal,
                 fg=self.fg_relatorio, font=("Segoe UI", 11, "bold")).pack(side="left")

        btn_restaurar_selecionados = tk.Button(header, text="Restaurar selecionados", bg=self.bg_nav_ativa,
                                               fg=self.fg_relatorio, activebackground=self.bg_card_hover,
                                               activeforeground=self.fg_relatorio, font=("Segoe UI", 9),
                                               relief="flat", padx=10, pady=4, bd=0, cursor="hand2",
                                               command=self._restaurar_ignorados_selecionados)
        btn_restaurar_selecionados.pack(side="right")

        btn_restaurar = tk.Button(header, text="Restaurar ignorados", bg=self.bg_nav_ativa,
                                  fg=self.fg_relatorio, activebackground=self.bg_card_hover,
                                  activeforeground=self.fg_relatorio, font=("Segoe UI", 9),
                                  relief="flat", padx=10, pady=4, bd=0, cursor="hand2",
                                  command=self._restaurar_ignorados)
        btn_restaurar.pack(side="right", padx=(0, 8))

        if total == 0:
            self.ignorados_relatorio_selecionados.clear()
            self.ignorados_relatorio_vars.clear()
            tk.Label(self.tab_ignorados._frame_conteudo, text="Nenhum arquivo ignorado.",
                     bg=self.bg_principal, fg=self.fg_texto_secundario,
                     font=("Segoe UI", 10)).pack(anchor="w", padx=10, pady=20)
            return

        self.ignorados_relatorio_selecionados = {
            chave for chave in self.ignorados_relatorio_selecionados
            if chave[0] in ignorados and chave[1] in ignorados[chave[0]]
        }
        self.ignorados_relatorio_vars.clear()

        for pasta, arquivos in sorted(ignorados.items(), key=lambda x: x[0].lower()):
            pasta_frame = tk.Frame(self.tab_ignorados._frame_conteudo, bg=self.bg_principal)
            pasta_frame.pack(fill="x", padx=10, pady=(10, 2))
            self._criar_label_selecionavel(
                pasta_frame,
                self._titulo_pasta(pasta),
                bg=self.bg_principal,
                fg=self.cor_desatualizado,
                font_spec=("Segoe UI", 10, "bold")
            )
            for nome_base in sorted(arquivos, key=str.lower):
                item_frame = tk.Frame(self.tab_ignorados._frame_conteudo, bg=self.bg_card)
                item_frame.pack(fill="x", padx=0, pady=(0, 8), ipady=9)
                chave = (pasta, nome_base)
                var = tk.BooleanVar(value=chave in self.ignorados_relatorio_selecionados)
                self.ignorados_relatorio_vars[chave] = var
                chk = tk.Checkbutton(
                    item_frame,
                    variable=var,
                    command=lambda p=pasta, n=nome_base, v=var: self._alternar_ignorado_relatorio(p, n, v),
                    bg=self.bg_card,
                    activebackground=self.bg_card,
                    selectcolor=self.bg_nav_ativa,
                    cursor="hand2",
                    bd=0,
                    highlightthickness=0
                )
                chk.pack(side="left", padx=(14, 4))
                tk.Label(item_frame, text=nome_base, bg=self.bg_card, fg=self.fg_relatorio,
                         font=("Segoe UI", 10, "bold"), anchor="w").pack(side="left", fill="x", expand=True, padx=(0, 18))

    def _alternar_ignorado_relatorio(self, pasta, nome_base, var):
        chave = (pasta, nome_base)
        if var.get():
            self.ignorados_relatorio_selecionados.add(chave)
        else:
            self.ignorados_relatorio_selecionados.discard(chave)

    def _restaurar_ignorados_selecionados(self):
        selecionados = sorted(self.ignorados_relatorio_selecionados, key=lambda item: (item[0].lower(), item[1].lower()))
        if not selecionados:
            messagebox.showinfo("Restaurar selecionados", "Selecione ao menos um arquivo ignorado.")
            return

        total = len(selecionados)
        if not messagebox.askyesno("Restaurar selecionados", f"Deseja restaurar {total} arquivo(s) selecionado(s)?"):
            return

        arquivos_ignorados = self.config.setdefault('arquivos_ignorados', {})
        for pasta, nome_base in selecionados:
            arquivos = arquivos_ignorados.get(pasta, [])
            if nome_base in arquivos:
                arquivos.remove(nome_base)
            if not arquivos and pasta in arquivos_ignorados:
                arquivos_ignorados.pop(pasta, None)

        self.ignorados_relatorio_selecionados.clear()
        self.ignorados_relatorio_vars.clear()
        salvar_config(self.config)
        self._preencher_aba_ignorados()
        messagebox.showinfo("Restaurar selecionados", "Arquivos selecionados restaurados.\nExecute a verificacao novamente.")

    def _restaurar_ignorados(self):
        total = sum(len(arquivos) for arquivos in self.config.get('arquivos_ignorados', {}).values())
        if total == 0:
            messagebox.showinfo("Restaurar ignorados", "Nenhum arquivo ignorado para restaurar.")
            return
        if not messagebox.askyesno("Restaurar ignorados", f"Deseja restaurar {total} arquivo(s) ignorado(s)?"):
            return

        self.config['arquivos_ignorados'] = {}
        self.ignorados_relatorio_selecionados.clear()
        self.ignorados_relatorio_vars.clear()
        salvar_config(self.config)
        self._preencher_aba_ignorados()
        messagebox.showinfo("Restaurar ignorados", "Arquivos ignorados restaurados.\nExecute a verificacao novamente.")

    def _preencher_aba_desatualizados(self, filtros):
        """Preenche a aba de desatualizados com dropdown para cada arquivo"""
        for widget in self.tab_desatualizados._frame_conteudo.winfo_children():
            widget.destroy()
        
        if not filtros['desatualizados']:
            lbl = tk.Label(self.tab_desatualizados._frame_conteudo, text="Filtro desativado", 
                          bg=self.bg_principal, fg=self.fg_texto, font=("Segoe UI", 9))
            lbl.pack(pady=20)
            return
        
        total_desatualizados = sum(len(r['desatualizados']) for r in self.resultados.values())
        
        if total_desatualizados == 0:
            lbl = tk.Label(self.tab_desatualizados._frame_conteudo, 
                          text="✅ Nenhum arquivo desatualizado\n(Tudo está sincronizado)", 
                          bg=self.bg_principal, fg=self.fg_texto, font=("Segoe UI", 10))
            lbl.pack(pady=20)
            return
        
        titulo_frame = tk.Frame(self.tab_desatualizados._frame_conteudo, bg=self.bg_principal)
        titulo_frame.pack(fill="x", padx=10, pady=(10, 5))
        tk.Label(titulo_frame, text=f"Total: {total_desatualizados} arquivo(s) desatualizado(s)",
                bg=self.bg_principal, fg=self.cor_desatualizado, font=("Segoe UI", 9, "bold")).pack(anchor="w")
        
        for caminho, resultado in sorted(self.resultados.items(), key=lambda x: x[0].lower()):
            desatualizados = resultado['desatualizados']
            if not desatualizados:
                continue
            
            pasta_frame = tk.Frame(self.tab_desatualizados._frame_conteudo, bg=self.bg_principal)
            pasta_frame.pack(fill="x", padx=10, pady=(10, 2))
            self._criar_label_selecionavel(pasta_frame, self._titulo_pasta(caminho), bg=self.bg_principal,
                    fg=self.cor_desatualizado, font_spec=("Segoe UI", 10, "bold"))
            
            config_frame = tk.Frame(self.tab_desatualizados._frame_conteudo, bg=self.bg_principal)
            config_frame.pack(fill="x", padx=20, pady=(0, 5))
            tk.Label(config_frame, text=f"Filtro: {resultado['config']['entrada'].upper()} → {resultado['config']['saida'].upper()}", 
                    bg=self.bg_principal, fg=self.fg_texto_secundario, font=("Segoe UI", 8)).pack(anchor="w")
            
            for item in desatualizados:
                nome_arquivo = f"{item['nome_base']} (há {item['dias']} dia(s))"
                caminho_arquivo = item['entrada']['caminho']
                nome_base = item['nome_base']
                self._adicionar_item_arquivo(self.tab_desatualizados._frame_conteudo, nome_arquivo, caminho, caminho_arquivo, nome_base)

    def _preencher_aba_atualizados(self, filtros):
        """Preenche a aba de atualizados com dropdown para cada arquivo"""
        for widget in self.tab_atualizados._frame_conteudo.winfo_children():
            widget.destroy()
        
        if not filtros['atualizados']:
            lbl = tk.Label(self.tab_atualizados._frame_conteudo, text="Filtro desativado", 
                          bg=self.bg_principal, fg=self.fg_texto, font=("Segoe UI", 9))
            lbl.pack(pady=20)
            return
        
        total_atualizados = sum(len(r['atualizados']) for r in self.resultados.values())
        
        if total_atualizados == 0:
            lbl = tk.Label(self.tab_atualizados._frame_conteudo, 
                          text="ℹ️ Nenhum arquivo atualizado\n(Execute a verificação para ver resultados)", 
                          bg=self.bg_principal, fg=self.fg_texto, font=("Segoe UI", 10))
            lbl.pack(pady=20)
            return
        
        titulo_frame = tk.Frame(self.tab_atualizados._frame_conteudo, bg=self.bg_principal)
        titulo_frame.pack(fill="x", padx=10, pady=(10, 5))
        tk.Label(titulo_frame, text=f"Total: {total_atualizados} arquivo(s) atualizado(s)",
                bg=self.bg_principal, fg=self.cor_atualizado, font=("Segoe UI", 9, "bold")).pack(anchor="w")
        
        for caminho, resultado in sorted(self.resultados.items(), key=lambda x: x[0].lower()):
            atualizados = resultado['atualizados']
            if not atualizados:
                continue
            
            pasta_frame = tk.Frame(self.tab_atualizados._frame_conteudo, bg=self.bg_principal)
            pasta_frame.pack(fill="x", padx=10, pady=(10, 2))
            self._criar_label_selecionavel(pasta_frame, self._titulo_pasta(caminho), bg=self.bg_principal,
                    fg=self.cor_atualizado, font_spec=("Segoe UI", 10, "bold"))
            
            config_frame = tk.Frame(self.tab_atualizados._frame_conteudo, bg=self.bg_principal)
            config_frame.pack(fill="x", padx=20, pady=(0, 5))
            tk.Label(config_frame, text=f"Filtro: {resultado['config']['entrada'].upper()} → {resultado['config']['saida'].upper()}", 
                    bg=self.bg_principal, fg=self.fg_texto_secundario, font=("Segoe UI", 8)).pack(anchor="w")
            
            for item in atualizados:
                nome_arquivo = f"{item['nome_base']} ({item['entrada']['data']})"
                caminho_arquivo = item['entrada']['caminho']
                nome_base = item['nome_base']
                self._adicionar_item_arquivo(self.tab_atualizados._frame_conteudo, nome_arquivo, caminho, caminho_arquivo, nome_base)

    def _exportar_csv(self):
        """Exporta o relatório para um arquivo Excel (.xlsx) com hiperlinks"""
        from datetime import datetime
        
        # Pedir caminho para salvar o arquivo em Excel
        caminho_arquivo = filedialog.asksaveasfilename(
            title="Salvar Relatório como Excel",
            defaultextension=".xlsx",
            filetypes=[("Arquivo Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
            initialfile=f"relatorio_{datetime.now().strftime('%d_%m_%Y_%H%M%S')}.xlsx"
        )
        
        if not caminho_arquivo:
            return
        
        try:
            # Garantir que seja .xlsx
            if not caminho_arquivo.endswith('.xlsx'):
                caminho_arquivo = caminho_arquivo.replace('.csv', '.xlsx').replace('.xls', '.xlsx')
                if not caminho_arquivo.endswith('.xlsx'):
                    caminho_arquivo += '.xlsx'
            
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Relatório"
            
            # Header com estilo
            headers = ['Modelo', 'Pasta', 'Status', 'Motivo']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="5865F2", end_color="5865F2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Coletar dados de todos os arquivos
            row_num = 2
            
            for caminho_pasta, resultado in sorted(self.resultados.items(), key=lambda x: x[0].lower()):
                for tipo_lista, tipo_nome in [('novos', 'Novo'), ('desatualizados', 'Desatualizado'), ('atualizados', 'Atualizado')]:
                    itens = resultado[tipo_lista]
                    
                    for item in itens:
                        nome_base = item.get('nome_base', '')
                        caminho_arquivo_item = item['entrada']['caminho']
                        pasta_arquivo = os.path.dirname(caminho_arquivo_item)  # Pasta onde o arquivo está
                        status = 'Novo' if tipo_nome == 'Novo' else 'Desatualizado' if tipo_nome == 'Desatualizado' else 'Atualizado'
                        
                        # Verificar se existe status salvo
                        if 'status_arquivos' in self.config and caminho_arquivo_item in self.config['status_arquivos']:
                            status = self.config['status_arquivos'][caminho_arquivo_item]
                        
                        # Excluir do relatório arquivos que estão na aba de atualizados
                        # sem status definido manualmente pelo usuário
                        if tipo_lista == 'atualizados' and caminho_arquivo_item not in self.config.get('status_arquivos', {}):
                            continue
                        
                        # Se status for "Status" (padrão), marcar como "Inalterado"
                        if status == "Status":
                            status = "Inalterado"
                        
                        # Obter motivo (relatório) se existir
                        motivo = ''
                        if 'relatorios_inaptid' in self.config and caminho_arquivo_item in self.config['relatorios_inaptid']:
                            motivo = self.config['relatorios_inaptid'][caminho_arquivo_item]
                        
                        # Modelo (coluna A)
                        cell_modelo = ws.cell(row=row_num, column=1)
                        cell_modelo.value = nome_base
                        cell_modelo.alignment = Alignment(horizontal="left", vertical="center")
                        
                        # Pasta como hiperlink (coluna B)
                        cell_pasta = ws.cell(row=row_num, column=2)
                        cell_pasta.value = os.path.basename(pasta_arquivo)  # Mostrar apenas o nome da pasta
                        
                        # Criar hiperlink para abrir a pasta
                        if os.path.exists(pasta_arquivo):
                            # Converter caminho para file:/// URI
                            file_uri = 'file:///' + pasta_arquivo.replace('\\', '/').replace(' ', '%20')
                            cell_pasta.hyperlink = file_uri
                            cell_pasta.font = Font(underline="single", color="0563C1")
                        else:
                            cell_pasta.font = Font(color="808080")
                        
                        cell_pasta.alignment = Alignment(horizontal="left", vertical="center")
                        
                        # Status (coluna C)
                        cell_status = ws.cell(row=row_num, column=3)
                        cell_status.value = status
                        cell_status.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Motivo (coluna D, com wrap de texto)
                        cell_motivo = ws.cell(row=row_num, column=4)
                        cell_motivo.value = motivo
                        cell_motivo.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                        
                        row_num += 1
            
            # Ajustar largura das colunas
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 60
            
            # Altura adequada para header
            ws.row_dimensions[1].height = 25
            
            # Auto-ajustar altura das linhas com conteúdo
            for row in range(2, row_num):
                ws.row_dimensions[row].height = None
            
            # Salvar arquivo
            wb.save(caminho_arquivo)
            
            messagebox.showinfo("Sucesso", f"✓ Relatório exportado com sucesso!\n\n{caminho_arquivo}")
            
        except ImportError:
            messagebox.showerror("Erro", "Para exportar em Excel, instale o pacote openpyxl:\n\npip install openpyxl")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar relatório:\n{str(e)}")
    
    def _ao_fechar(self):
        """Recarrega config da aplicação pai ao fechar a janela"""
        self.parent.config = carregar_config()
        if self.on_voltar:
            self.on_voltar()
            return
        self.destroy()


# ==================== INTERFACE PRINCIPAL ====================

class MonitorApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("ATLAS")
        self.geometry("500x230")
        self.resizable(False, False)
        
        self.bg_principal = "#36393F"
        self.bg_secundario = "#2F3136"
        self.bg_terciario = "#282B30"
        self.fg_texto = "#FFFFFF"
        self.fg_texto_secundario = "#B9BBBE"
        self.cor_acento = "#5865F2"
        self.cor_sucesso = "#43B581"
        self.cor_alerta = "#FAA61A"
        self.cor_erro = "#F04747"
        
        self.configure(bg=self.bg_principal)
        
        self._configurar_estilo_ttk()

        self.config = carregar_config()
        self.sessao_atual = self.config['sessao_ativa']
        self.pastas = obter_pastas_sessao(self.config, self.sessao_atual)
        self.resultados = {}

        ult = self.config.get('ultimas_extensoes', {})
        self.ultimo_entrada = ult.get('entrada', 'rvt')
        self.ultimo_saida = ult.get('saida', 'ifc')

        self.filter_novos = tk.BooleanVar(value=True)
        self.filter_desatualizados = tk.BooleanVar(value=True)
        self.filter_atualizados = tk.BooleanVar(value=True)
        
        self.visualizando_ignorados = False

        self._montar_interface()
        
        # Configurar handler para fechar a janela
        self.protocol("WM_DELETE_WINDOW", self._ao_fechar_aplicacao)
        
        self.log_console(f"📁 Configurações salvas em: {DIRETORIO_ATLAS}")
        self.log_console(f"📄 Arquivo: {os.path.basename(CONFIG_FILE)}\n")
    
    def _configurar_estilo_ttk(self):
        """Configura o estilo visual dos widgets ttk"""
        style = ttk.Style()
        style.theme_use('clam')
        
        style.configure('TLabel', background=self.bg_principal, foreground=self.fg_texto)
        style.configure('TFrame', background=self.bg_principal)
        style.configure('TButton', background=self.bg_terciario, foreground=self.fg_texto, 
                       borderwidth=1, focuscolor='none', padding=6)
        style.map('TButton', 
                 background=[('active', self.cor_acento), ('pressed', self.cor_acento)],
                 foreground=[('active', '#ffffff'), ('pressed', '#ffffff')])
        
        style.configure('TCombobox', fieldbackground="#5865F2", background="#5865F2",
                       foreground=self.fg_texto)
        style.map('TCombobox', fieldbackground=[('readonly', '#5865F2'), ('active', '#5865F2')])
        
        style.configure('TCheckbutton', background=self.bg_principal, foreground=self.fg_texto)
        style.map('TCheckbutton', background=[('active', self.bg_principal)])

    def _montar_interface(self):
        self.tela_principal = tk.Frame(self, bg=self.bg_principal)
        self.tela_principal.pack(fill="both", expand=True)
        self.tela_relatorio = None

        header_frame = tk.Frame(self.tela_principal, bg=self.bg_secundario, height=60)
        header_frame.pack(fill="x", pady=(0, 10))
        header_frame.pack_propagate(False)
        
        ttk.Label(
            header_frame,
            text="🔄 Monitor de Atualizações de Arquivos",
            font=("Segoe UI", 16, "bold"),
            background=self.bg_secundario,
            foreground=self.cor_acento,
        ).pack(pady=12)

        frame_sessao = tk.Frame(self.tela_principal, bg=self.bg_principal)
        frame_sessao.pack(pady=8)
        
        ttk.Label(frame_sessao, text="Sessão Ativa:", font=("Segoe UI", 10, "bold"), background=self.bg_principal, foreground=self.fg_texto).pack(side="left", padx=5)
        
        self.combo_sessoes = ttk.Combobox(frame_sessao, width=30, state="readonly")
        self.combo_sessoes.pack(side="left", padx=5)
        self.combo_sessoes.bind("<<ComboboxSelected>>", self._trocar_sessao)
        self._atualizar_combo_sessoes()
        
        ttk.Button(frame_sessao, text="⚙️ Gerenciar Sessões", command=self._gerenciar_sessoes, width=20).pack(side="left", padx=5)
        
        frame_botoes = tk.Frame(self.tela_principal, bg=self.bg_principal)
        frame_botoes.pack(pady=8)

        ttk.Button(
            frame_botoes, text="📂 Adicionar Pasta(s)", width=25, command=self._adicionar_pastas
        ).grid(row=0, column=0, padx=8, pady=4)
        ttk.Button(
            frame_botoes, text="🔍 Verificar Atualizações", width=25, command=self._verificar_atualizacoes
        ).grid(row=0, column=1, padx=8, pady=4)
        ttk.Button(
            frame_botoes, text="📁 Mapeamento", width=25, command=self._gerenciar_pastas
        ).grid(row=1, column=0, padx=8, pady=4)
        ttk.Button(
            frame_botoes, text="📊 Relatório", width=25, command=self._abrir_relatorio
        ).grid(row=1, column=1, padx=8, pady=4)
    def _atualizar_combo_sessoes(self):
        """Atualiza o combobox com as sessões disponíveis"""
        sessoes = list(self.config['sessoes'].keys())
        self.combo_sessoes['values'] = sessoes
        self.combo_sessoes.set(self.sessao_atual)

    def _trocar_sessao(self, event=None):
        """Troca para a sessão selecionada"""
        nova_sessao = self.combo_sessoes.get()
        if nova_sessao == self.sessao_atual:
            return
        
        self.sessao_atual = nova_sessao
        self.config['sessao_ativa'] = nova_sessao
        salvar_config(self.config)
        
        self.carregar_sessao_ativa()
        self.log_console(f"✓ Sessão alterada para: {nova_sessao}")
        self.log_console(f"📁 Configurações salvas em: {DIRETORIO_ATLAS}\n")

    def carregar_sessao_ativa(self):
        """Carrega as pastas da sessão ativa"""
        self.pastas = obter_pastas_sessao(self.config, self.sessao_atual)
        self._atualizar_combo_sessoes()

    def _gerenciar_sessoes(self):
        """Abre janela de gerenciamento de sessões"""
        JanelaGerenciarSessoes(self, self.config, self.bg_principal, self.bg_secundario, self.fg_texto, self.fg_texto_secundario)

    def _gerenciar_pastas(self):
        """Abre janela de gerenciamento de pastas"""
        JanelaGerenciarPastas(self, self.config, self.bg_principal, self.bg_secundario, self.fg_texto, self.fg_texto_secundario)

    def _abrir_relatorio(self):
        """Mostra o relatório dentro da janela principal"""
        # Recarregar config do disco para garantir dados atualizados
        self.config = carregar_config()
        self.tela_principal.pack_forget()
        self.geometry("700x680")
        self.minsize(900, 620)
        self.resizable(True, True)

        if self.tela_relatorio:
            self.tela_relatorio.destroy()

        self.tela_relatorio = JanelaRelatorio(
            self,
            self.config,
            self.resultados,
            self.bg_principal,
            self.bg_secundario,
            self.fg_texto,
            self.fg_texto_secundario,
            self.filter_novos,
            self.filter_desatualizados,
            self.filter_atualizados,
            on_voltar=self._voltar_tela_principal
        )
        self.tela_relatorio.pack(fill="both", expand=True)

    def _voltar_tela_principal(self):
        """Volta do relatório para a tela principal"""
        self.config = carregar_config()
        if self.tela_relatorio:
            self.tela_relatorio.destroy()
            self.tela_relatorio = None
        self.tela_principal.pack(fill="both", expand=True)
        self.minsize(500, 230)
        self.geometry("500x230")
        self.resizable(False, False)

    def log_console(self, msg: str):
        """Log de mensagens no console"""
        print(msg)

    def _ao_fechar_aplicacao(self):
        """Handler para quando a janela principal é fechada"""
        if messagebox.askyesno("Confirmar Saída", "Tem certeza que deseja fechar o ATLAS?"):
            self.destroy()

    def rodar_em_thread(self, func):
        threading.Thread(target=func, daemon=True).start()

    def _ignorar_arquivos(self):
        """Permite selecionar arquivos para ignorar da pasta selecionada"""
        pass

    def _reverter_ignorados(self):
        """Remove todos os arquivos ignorados da pasta selecionada"""
        pass

    def _verificar_atualizacoes(self):
        if not self.pastas:
            messagebox.showinfo("Aviso", "Nenhuma pasta configurada nesta sessão.")
            return

        self.visualizando_ignorados = False

        dialog = tk.Toplevel(self)
        dialog.title("Verificação Concluída")
        dialog.geometry("280x160")
        dialog.resizable(False, False)
        dialog.grab_set()
        dialog.transient(self)
        dialog.configure(bg=self.bg_principal)
        
        header_frame = tk.Frame(dialog, bg=self.bg_principal)
        header_frame.pack(pady=5, padx=15)
        
        tk.Label(
            header_frame,
            text="✓ Verificação Concluída",
            font=("Segoe UI", 10, "bold"),
            background=self.bg_principal,
            foreground=self.cor_acento
        ).pack(anchor="w")
        
        info_frame = tk.Frame(dialog, bg=self.bg_principal)
        info_frame.pack(pady=(0, 5), padx=15, fill="both", expand=True)
        
        info_labels = {
            'novos': tk.Label(
                info_frame,
                text="🆕 Novos: --",
                font=("Segoe UI", 9),
                background=self.bg_principal,
                foreground=self.fg_texto
            ),
            'desatualizados': tk.Label(
                info_frame,
                text="⚠️ Desatualizados: --",
                font=("Segoe UI", 9),
                background=self.bg_principal,
                foreground=self.fg_texto
            ),
            'atualizados': tk.Label(
                info_frame,
                text="✅ Atualizados: --",
                font=("Segoe UI", 9),
                background=self.bg_principal,
                foreground=self.fg_texto
            )
        }
        
        tk.Label(
            info_frame,
            text="Resultados da Verificação:",
            font=("Segoe UI", 8),
            background=self.bg_principal,
            foreground=self.fg_texto
        ).pack(anchor="center", pady=(0, 2))
        
        for label in info_labels.values():
            label.pack(anchor="center", pady=0)
        
        button_frame = tk.Frame(dialog, bg=self.bg_principal)
        button_frame.pack(pady=5, fill="x", padx=15)
        
        buttons_inner = tk.Frame(button_frame, bg=self.bg_principal)
        buttons_inner.pack(anchor="center")
        
        def tarefa_verificacao():
            # Resetar TODOS os status e relatórios para nova verificação limpa
            self.config['status_arquivos'] = {}
            self.config['relatorios_inaptid'] = {}
            salvar_config(self.config)
            
            self.resultados = verificar_atualizacoes(self.pastas)
            self.resultados = self._filtrar_ignorados(self.resultados)

            total_novos = sum(len(r['novos']) for r in self.resultados.values())
            total_desatualizados = sum(len(r['desatualizados']) for r in self.resultados.values())
            total_atualizados = sum(len(r['atualizados']) for r in self.resultados.values())

            info_labels['novos'].config(text=f"🆕 Novos: {total_novos}")
            info_labels['desatualizados'].config(text=f"⚠️ Desatualizados: {total_desatualizados}")
            info_labels['atualizados'].config(text=f"✅ Atualizados: {total_atualizados}")
        
        def abrir_relatorio():
            dialog.destroy()
            self._abrir_relatorio()
        
        def atualizar():
            dialog.destroy()
            
            temp_file, command_file = set_module.gerar_temp_set(self.resultados)
            
            if not temp_file:
                messagebox.showwarning("Aviso", "Nenhum arquivo novo ou desatualizado para processar.")
                return
            
            versoes_revit = detectar_versoes_revit()
            
            if not versoes_revit:
                messagebox.showerror("Revit Não Encontrado", "Nenhuma versão de Revit foi detectada no sistema.")
                return
            
            janela_revit = JanelaSelecaoRevit(self, versoes_revit)
            self.wait_window(janela_revit)
            
            if not janela_revit.revit_selecionado:
                return
            
            nome_revit, caminho_revit = janela_revit.revit_selecionado
            
            tela = tela_carregamento(
                titulo="Carregando Revit",
                mensagem=f"Abrindo {nome_revit}...",
                duracao=60,
                parent=self
            )
            
            resultado_abertura = abrir_revit(caminho_revit)
        
        def gerar_relatorio():
            dialog.destroy()
            destino = set_module.gerar_set(self.resultados, parent=self)
            if destino:
                messagebox.showinfo("Relatório Gerado", f"Relatório salvo em:\n{destino}")
        
        ttk.Button(
            buttons_inner,
            text="Atualizar",
            command=atualizar,
            width=11
        ).pack(side="left", padx=2)
        
        ttk.Button(
            buttons_inner,
            text="Gerar set",
            command=gerar_relatorio,
            width=11
        ).pack(side="left", padx=2)
        
        ttk.Button(
            buttons_inner,
            text="Relatório",
            command=abrir_relatorio,
            width=11
        ).pack(side="left", padx=2)
        
        self.rodar_em_thread(tarefa_verificacao)
        self.wait_window(dialog)

    def _adicionar_pastas(self):
        """
        Abre a janela de seleção múltipla. Para cada pasta confirmada,
        abre o diálogo de escolha de extensões usando como padrão as
        últimas escolhas (persistidas em config). 
        """
        janela = JanelaSelecaoPastas(self)
        self.wait_window(janela)

        if not getattr(janela, 'pastas_selecionadas', None):
            return

        for caminho in janela.pastas_selecionadas:
            existe = any(p['caminho'] == caminho for p in self.pastas)
            if existe:
                self.log_console(f"⚠️ Pasta já configurada, pulando: {caminho}")
                continue

            self._configurar_extensoes_pasta(caminho)

    def _configurar_extensoes_pasta(self, caminho):
        entrada_win = tk.Toplevel(self)
        entrada_win.title("Escolher Extensões")
        entrada_win.geometry("420x400")
        entrada_win.resizable(False, False)
        entrada_win.grab_set()
        entrada_win.transient(self)
        entrada_win.configure(bg="#f0f0f0")

        confirmado = [False]

        main_frame = ttk.Frame(entrada_win, padding="20 15 20 20")
        main_frame.pack(fill="both", expand=True)

        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill="x", pady=(0, 15))

        ttk.Label(
            header_frame,
            text="📁 Configurar Extensões",
            font=("Segoe UI", 14, "bold"),
        ).pack(anchor="w")

        ttk.Separator(main_frame).pack(fill="x", pady=(0, 15))

        info_frame = ttk.Frame(main_frame)
        info_frame.pack(fill="x", pady=(0, 15))

        ttk.Label(
            info_frame,
            text="Pasta Selecionada:",
            font=("Segoe UI", 10, "bold"),
        ).pack(anchor="w")

        ttk.Label(
            info_frame,
            text=os.path.basename(caminho),
            font=("Segoe UI", 11),
            foreground="#1a73e8",
        ).pack(anchor="w", pady=(2, 5))

        ttk.Label(
            info_frame,
            text="Caminho completo:",
            font=("Segoe UI", 9),
        ).pack(anchor="w")

        caminho_label = ttk.Label(
            info_frame,
            text=caminho,
            font=("Segoe UI", 9),
            foreground="#666666",
            wraplength=380
        )
        caminho_label.pack(anchor="w", pady=(0, 10))

        ext_frame = ttk.Frame(main_frame)
        ext_frame.pack(fill="x", pady=(0, 20))

        entrada_frame = ttk.Frame(ext_frame)
        entrada_frame.pack(fill="x", pady=(0, 15))

        ttk.Label(
            entrada_frame,
            text="Extensão de Entrada:",
            font=("Segoe UI", 10),
        ).pack(anchor="w", pady=(0, 5))

        default_entrada = self.config.get('ultimas_extensoes', {}).get('entrada', self.ultimo_entrada)
        entrada_var = tk.StringVar(value=default_entrada)
        entrada_combo = ttk.Combobox(
            entrada_frame,
            textvariable=entrada_var,
            values=["dwg", "rvt", "rfa"],
            state="readonly",
            width=25
        )
        entrada_combo.pack(anchor="w")

        saida_frame = ttk.Frame(ext_frame)
        saida_frame.pack(fill="x")

        ttk.Label(
            saida_frame,
            text="Extensão de Saída:",
            font=("Segoe UI", 10),
        ).pack(anchor="w", pady=(0, 10))

        default_saida = self.config.get('ultimas_extensoes', {}).get('saida', self.ultimo_saida)
        saida_var = tk.StringVar(value=default_saida)
        saida_combo = ttk.Combobox(
            saida_frame,
            textvariable=saida_var,
            values=["ifc", "nwc"],
            state="readonly",
            width=25
        )
        saida_combo.pack(anchor="w")

        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=(10, 0))

        ttk.Button(
            button_frame,
            text="Cancelar",
            command=entrada_win.destroy,
            style="Secondary.TButton",
            width=15
        ).pack(side="right", padx=(5, 0))

        def confirmar():
            self.pastas.append({
                "caminho": caminho,
                "entrada": entrada_var.get(),
                "saida": saida_var.get()
            })
            self.ultimo_entrada = entrada_var.get()
            self.ultimo_saida = saida_var.get()
            self.config['ultimas_extensoes'] = {
                'entrada': self.ultimo_entrada,
                'saida': self.ultimo_saida
            }
            atualizar_pastas_sessao(self.config, self.sessao_atual, self.pastas)
            salvar_config(self.config)
            self.log_console(f"📂 Pasta adicionada: {caminho} ({entrada_var.get()} → {saida_var.get()})")
            confirmado[0] = True
            entrada_win.destroy()

        ttk.Button(
            button_frame,
            text="✓ Confirmar",
            command=confirmar,
            style="Primary.TButton",
            width=15
        ).pack(side="right", padx=(0, 5))

        style = ttk.Style()
        style.configure("Primary.TButton", background="#007bff", foreground="black")
        style.configure("Secondary.TButton", background="#6c757d")

        entrada_win.update_idletasks()
        width = entrada_win.winfo_width()
        height = entrada_win.winfo_height()
        x = (entrada_win.winfo_screenwidth() // 2) - (width // 2)
        y = (entrada_win.winfo_screenheight() // 2) - (height // 2)
        entrada_win.geometry(f"{width}x{height}+{x}+{y}")

        self.wait_window(entrada_win)
        return confirmado[0]

    def _filtrar_ignorados(self, resultados):
        """Filtra resultados removendo arquivos ignorados"""
        if 'arquivos_ignorados' not in self.config:
            return resultados
        
        resultados_filtrados = {}
        for caminho_pasta, resultado in resultados.items():
            ignorados = set(self.config['arquivos_ignorados'].get(caminho_pasta, []))
            
            novos = [r for r in resultado['novos'] if not self._arquivo_esta_ignorado(r, ignorados)]
            desatualizados = [r for r in resultado['desatualizados'] if not self._arquivo_esta_ignorado(r, ignorados)]
            atualizados = [r for r in resultado['atualizados'] if not self._arquivo_esta_ignorado(r, ignorados)]
            
            resultados_filtrados[caminho_pasta] = {
                'config': resultado['config'],
                'novos': novos,
                'desatualizados': desatualizados,
                'atualizados': atualizados,
                'total_entrada': resultado['total_entrada'],
                'total_saida': resultado['total_saida']
            }
        
        return resultados_filtrados

    def _arquivo_esta_ignorado(self, arquivo_info, ignorados):
        """Verifica se um arquivo está na lista de ignorados"""
        nome_base = arquivo_info.get('nome_base', '')
        
        if nome_base in ignorados:
            return True
        
        info_entrada = arquivo_info.get('entrada', {})
        if info_entrada:
            arquivo_entrada = info_entrada.get('arquivo', '')
            if arquivo_entrada in ignorados:
                return True
        
        nome_sem_ext = nome_base.rsplit('.', 1)[0] if '.' in nome_base else nome_base
        for ignorado in ignorados:
            ignorado_sem_ext = ignorado.rsplit('.', 1)[0] if '.' in ignorado else ignorado
            if nome_sem_ext == ignorado_sem_ext:
                return True
        
        return False


if __name__ == "__main__":
    app = MonitorApp()
    app.mainloop()
